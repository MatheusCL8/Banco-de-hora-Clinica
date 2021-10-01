from util import sleep
from openpyxl import *
from itertools import islice
from openpyxl.utils.dataframe import dataframe_to_rows

def data_e_hora():
    from datetime import date,datetime
    hoje = date.today()
    _hoje=str(hoje).split('-')
    _hoje=_hoje[2]+'/'+_hoje[1]+'/'+_hoje[0]

    hd=datetime.now()
    horario_atual=hd.strftime('%H:%M:%S')
    return _hoje,horario_atual,hoje

def criar_folha_func(nome,excel,lista_func,caminho):
    excel.create_sheet(title=nome)
    dados_func=excel[nome]
    dados_func.cell(row=1,column=1,value='DATA')
    dados_func.cell(row=1,column=2,value='ENTRADA_MANHA')
    dados_func.cell(row=1,column=3,value='SAIDA_MANHA')
    dados_func.cell(row=1,column=4,value='ENTRADA_TARDE')
    dados_func.cell(row=1,column=5,value='SAIDA_TARDE')
    l=1
    for item in lista_func['A']:
        l+=1
    lista_func.cell(row=l,column=1,value=nome)
    lista_func.cell(row=l,column=2,value=0)
    excel.save(filename=caminho)

def folha_de_ponto(nome_funcionario):
    nome=nome_funcionario
    data,hora,data_comparar=data_e_hora()
    caminho='./_excel/banco_de_horas.xlsx'
    wb=load_workbook(filename=caminho)
    cont=wb['FUNCIONARIOS']

    validador=''
    
    try:
        dados_func=wb[nome_funcionario]
        contador=''
        valor_linha=''
        for item in cont['A']:
            if item.value ==nome_funcionario:
                valor_linha=item.row
                contador=cont[f'B{item.row}'].value
        linha=1

        for item in dados_func['A']:
            linha+=1

        if contador==0:
            dados_func.cell(row=linha,column=1,value=data)
            dados_func.cell(row=linha,column=2,value=hora)
            cont.cell(row=valor_linha,column=2,value=1)
            wb.save(filename=caminho)
            validador=True
        
        elif contador==0 and data_comparar.weekday()==5:
            dados_func.cell(row=linha,column=1,value=data)
            dados_func.cell(row=linha,column=2,value=hora)
            cont.cell(row=valor_linha,column=2,value=1)
            wb.save(filename=caminho)
            validador=True
            
        elif contador==1:
            dados_func.cell(row=linha-1,column=3,value=hora)
            cont.cell(row=valor_linha,column=2,value=2)
            wb.save(filename=caminho)
            validador=True
        
        elif contador==1 and data_comparar.weekday()==5:
            dados_func.cell(row=linha-1,column=3,value=hora)
            cont.cell(row=valor_linha,column=2,value=0)
            wb.save(filename=caminho)
            validador=True

        elif contador==2:
            dados_func.cell(row=linha-1,column=4,value=hora)
            cont.cell(row=valor_linha,column=2,value=3)
            wb.save(filename=caminho)
            validador=True
        
        elif contador==3:
            dados_func.cell(row=linha-1,column=5,value=hora)
            cont.cell(row=valor_linha,column=2,value=0)
            wb.save(filename=caminho)
            validador=True

        return validador
    
    except:
        criar_folha_func(nome_funcionario,wb,cont,caminho)
        validador=False
        return validador
        

